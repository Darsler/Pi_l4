import re
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import hashlib

# Константы для работы с Excel
EXCEL_FILE = "restaurant_data.xlsx"
SHEET_USERS = "Пользователи"
SHEET_RESTAURANTS = "Рестораны"
SHEET_MENU = "Меню"


class Product:
    def __init__(self, id_, restaurant_id, name, price, status=True):
        self.id = id_
        self.restaurant_id = restaurant_id
        self.name = name
        self.price = float(price)
        self.status = bool(status)

    def change_status(self):
        self.status = not self.status
        return self.status

    def update_price(self, new_price):
        self.price = float(new_price)

    def update_name(self, new_name):
        self.name = new_name

    def __str__(self):
        return f"{self.name} - {self.price}₽ ({'доступен' if self.status else 'не доступен'})"


class Restaurant:
    phone_pattern = r'^(8\d{10}|\+7\d{10}|\d{10})$'

    def __init__(self, id_, name, phone, address):
        self.id = id_
        self.name = name
        self.phone = self.phone_check(phone)
        self.address = address

    def phone_check(self, phone_raw):
        if re.fullmatch(self.phone_pattern, phone_raw):
            return phone_raw
        else:
            raise ValueError("Некорректный номер телефона")

    def update_info(self, new_name=None, new_phone=None, new_address=None):
        if new_name:
            self.name = new_name
        if new_phone:
            self.phone = self.phone_check(new_phone)
        if new_address:
            self.address = new_address

    def __str__(self):
        return f"Ресторан: {self.name}\nАдрес: {self.address}\nТелефон: {self.phone}"


class ExcelManager:
    def __init__(self):
        self.file = EXCEL_FILE
        self._init_excel_file()

    def _init_excel_file(self):
        if not os.path.exists(self.file):
            wb = Workbook()

            # Лист пользователей
            ws = wb.active
            ws.title = SHEET_USERS
            ws.append(["ID", "Логин", "Пароль", "Роль"])

            # Лист ресторанов
            wb.create_sheet(SHEET_RESTAURANTS)
            wb[SHEET_RESTAURANTS].append(["ID", "Название", "Телефон", "Адрес"])

            # Лист меню
            wb.create_sheet(SHEET_MENU)
            wb[SHEET_MENU].append(["ID", "ID_ресторана", "Название", "Цена", "Статус"])

            # Добавляем администратора по умолчанию
            admin_pass = self._hash_password("admin")
            ws.append([1, "admin", admin_pass, "admin"])

            wb.save(self.file)

    def _hash_password(self, password):
        return hashlib.sha256(password.encode()).hexdigest()

    def verify_user(self, username, password):
        wb = openpyxl.load_workbook(self.file)
        ws = wb[SHEET_USERS]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == username and row[2] == self._hash_password(password):
                return {"id": row[0], "role": row[3]}
        return None

    def get_restaurants(self):
        wb = openpyxl.load_workbook(self.file)
        ws = wb[SHEET_RESTAURANTS]

        restaurants = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Проверяем, что ID не пустой
                restaurants.append(Restaurant(row[0], row[1], row[2], row[3]))
        return restaurants

    def get_products_for_restaurant(self, restaurant_id):
        wb = openpyxl.load_workbook(self.file)
        ws = wb[SHEET_MENU]

        products = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == restaurant_id:
                products.append(Product(row[0], row[1], row[2], row[3], row[4]))
        return products

    def save_restaurant(self, restaurant):
        wb = openpyxl.load_workbook(self.file)
        ws = wb[SHEET_RESTAURANTS]

        # Находим максимальный ID
        max_id = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] and row[0] > max_id:
                max_id = row[0]

        # Если это новый ресторан (без ID)
        if not hasattr(restaurant, 'id') or not restaurant.id:
            restaurant.id = max_id + 1
            ws.append([restaurant.id, restaurant.name, restaurant.phone, restaurant.address])
        else:
            # Обновляем существующий ресторан
            for row in ws.iter_rows(min_row=2):
                if row[0].value == restaurant.id:
                    row[1].value = restaurant.name
                    row[2].value = restaurant.phone
                    row[3].value = restaurant.address
                    break

        wb.save(self.file)
        return restaurant

    def save_product(self, product):
        wb = openpyxl.load_workbook(self.file)
        ws = wb[SHEET_MENU]

        # Находим максимальный ID
        max_id = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] and row[0] > max_id:
                max_id = row[0]

        # Если это новое блюдо (без ID)
        if not hasattr(product, 'id') or not product.id:
            product.id = max_id + 1
            ws.append([product.id, product.restaurant_id, product.name, product.price, product.status])
        else:
            # Обновляем существующее блюдо
            for row in ws.iter_rows(min_row=2):
                if row[0].value == product.id:
                    row[2].value = product.name
                    row[3].value = product.price
                    row[4].value = product.status
                    break

        wb.save(self.file)
        return product

    def delete_restaurant(self, restaurant_id):
        wb = openpyxl.load_workbook(self.file)

        # Удаляем ресторан
        ws = wb[SHEET_RESTAURANTS]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == restaurant_id:
                ws.delete_rows(row[0].row)
                break

        # Удаляем блюда этого ресторана
        ws = wb[SHEET_MENU]
        rows_to_delete = []
        for row in ws.iter_rows(min_row=2):
            if row[1].value == restaurant_id:
                rows_to_delete.append(row[0].row)

        for row_idx in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(row_idx)

        wb.save(self.file)

    def delete_product(self, product_id):
        wb = openpyxl.load_workbook(self.file)
        ws = wb[SHEET_MENU]

        for row in ws.iter_rows(min_row=2):
            if row[0].value == product_id:
                ws.delete_rows(row[0].row)
                break

        wb.save(self.file)


class RestaurantManager:
    def __init__(self):
        self.excel = ExcelManager()
        self.current_user = None
        self.restaurants = []
        self.load_data()

    def load_data(self):
        self.restaurants = self.excel.get_restaurants()
        for restaurant in self.restaurants:
            restaurant.menu = self.excel.get_products_for_restaurant(restaurant.id)

    def login(self):
        print("\nВход в систему")
        username = input("Логин: ")
        password = input("Пароль: ")

        user = self.excel.verify_user(username, password)
        if user:
            self.current_user = user
            print(f"\nДобро пожаловать, {username}!")
            return True
        else:
            print("Неверный логин или пароль")
            return False

    def create_restaurant(self):
        name_raw = input("Название ресторана: ")
        address_raw = input("Адрес: ")
        while True:
            phone_raw = input("Номер телефона (форматы: 8ХХХХХХХХХХ, +7ХХХХХХХХХХ): ")
            try:
                restaurant = Restaurant(None, name_raw, phone_raw, address_raw)
                restaurant = self.excel.save_restaurant(restaurant)
                restaurant.menu = []
                self.restaurants.append(restaurant)
                print("Ресторан успешно добавлен!")
                return
            except ValueError as e:
                print(f"Ошибка: {e}. Попробуйте еще раз.")

    def create_product(self, restaurant):
        name = input("Введите название блюда: ")
        while True:
            price = input("Введите цену блюда: ")
            try:
                product = Product(None, restaurant.id, name, float(price))
                product = self.excel.save_product(product)
                restaurant.menu.append(product)
                print("Блюдо успешно добавлено!")
                return
            except ValueError:
                print("Цена должна быть числом. Попробуйте еще раз.")

    def select_restaurant(self, prompt="Выберите ресторан: "):
        if not self.restaurants:
            print("Нет доступных ресторанов")
            return None

        print("\nСписок ресторанов:")
        for i, r in enumerate(self.restaurants, 1):
            print(f"{i}. {r.name} - {r.address}")

        try:
            choice = int(input(prompt)) - 1
            if 0 <= choice < len(self.restaurants):
                return self.restaurants[choice]
            else:
                print("Неверный номер ресторана")
                return None
        except ValueError:
            print("Введите число!")
            return None

    def show_all_restaurants(self):
        if not self.restaurants:
            print("Нет доступных ресторанов")
        else:
            print("\nСписок ресторанов:")
            for i, restaurant in enumerate(self.restaurants, 1):
                print(f"{i}. {restaurant.name} - {restaurant.address}")

    def show_restaurant_details(self):
        if not self.restaurants:
            print("Нет доступных ресторанов")
            return

        self.show_all_restaurants()
        try:
            choice = int(input("Введите номер ресторана для просмотра: ")) - 1
            if 0 <= choice < len(self.restaurants):
                print("\n" + str(self.restaurants[choice]))
                self.show_menu(self.restaurants[choice])
            else:
                print("Неверный номер ресторана")
        except ValueError:
            print("Введите число!")

    def search_restaurant(self):
        query = input("Введите часть названия или адреса для поиска: ").lower()
        found = [r for r in self.restaurants if query in r.name.lower() or query in r.address.lower()]

        if not found:
            print("Ничего не найдено")
        else:
            print("\nРезультаты поиска:")
            for i, restaurant in enumerate(found, 1):
                print(f"{i}. {restaurant.name} - {restaurant.address}")

    def show_menu(self, restaurant):
        if not restaurant.menu:
            print("Меню пустое")
            return False
        else:
            print(f"\nМеню ресторана '{restaurant.name}':")
            for i, product in enumerate(restaurant.menu, 1):
                print(f"{i}. {product}")
            return True

    def select_product(self, restaurant, prompt="Выберите блюдо: "):
        if not self.show_menu(restaurant):
            return None

        try:
            choice = int(input(prompt)) - 1
            if 0 <= choice < len(restaurant.menu):
                return restaurant.menu[choice]
            else:
                print("Неверный номер блюда")
                return None
        except ValueError:
            print("Введите число!")
            return None

    def edit_restaurant(self):
        restaurant = self.select_restaurant("Выберите ресторан для редактирования: ")
        if not restaurant:
            return

        print("\nТекущая информация:")
        print(restaurant)

        print("\nЧто вы хотите изменить?")
        print("1. Название")
        print("2. Телефон")
        print("3. Адрес")
        print("4. Удалить ресторан")
        print("5. Вернуться назад")

        choice = input("Выберите действие: ")

        if choice == "1":
            new_name = input("Введите новое название: ")
            restaurant.update_info(new_name=new_name)
            self.excel.save_restaurant(restaurant)
            print("Название успешно изменено!")
        elif choice == "2":
            while True:
                new_phone = input("Введите новый телефон: ")
                try:
                    restaurant.update_info(new_phone=new_phone)
                    self.excel.save_restaurant(restaurant)
                    print("Телефон успешно изменён!")
                    break
                except ValueError as e:
                    print(f"Ошибка: {e}")
        elif choice == "3":
            new_address = input("Введите новый адрес: ")
            restaurant.update_info(new_address=new_address)
            self.excel.save_restaurant(restaurant)
            print("Адрес успешно изменён!")
        elif choice == "4":
            confirm = input(f"Вы уверены, что хотите удалить ресторан '{restaurant.name}'? (да/нет): ")
            if confirm.lower() == 'да':
                self.excel.delete_restaurant(restaurant.id)
                self.restaurants.remove(restaurant)
                print("Ресторан успешно удалён!")
        elif choice == "5":
            return
        else:
            print("Неверный ввод")

    def edit_menu(self):
        restaurant = self.select_restaurant("Выберите ресторан для редактирования меню: ")
        if not restaurant:
            return

        while True:
            print("\nРедактирование меню:")
            if not self.show_menu(restaurant):
                print("1. Добавить блюдо")
                print("2. Вернуться назад")

                choice = input("Выберите действие: ")

                if choice == "1":
                    self.create_product(restaurant)
                elif choice == "2":
                    return
                else:
                    print("Неверный ввод")
            else:
                print("1. Добавить блюдо")
                print("2. Изменить блюдо")
                print("3. Удалить блюдо")
                print("4. Изменить статус блюда")
                print("5. Вернуться назад")

                choice = input("Выберите действие: ")

                if choice == "1":
                    self.create_product(restaurant)
                elif choice == "2":
                    product = self.select_product(restaurant, "Выберите блюдо для изменения: ")
                    if product:
                        new_name = input(f"Введите новое название (текущее: {product.name}): ") or product.name
                        new_price = input(f"Введите новую цену (текущая: {product.price}): ") or product.price

                        product.update_name(new_name)
                        try:
                            product.update_price(new_price)
                            self.excel.save_product(product)
                            print("Блюдо успешно изменено!")
                        except ValueError:
                            print("Цена должна быть числом. Изменения не сохранены.")
                elif choice == "3":
                    product = self.select_product(restaurant, "Выберите блюдо для удаления: ")
                    if product:
                        confirm = input(f"Вы уверены, что хотите удалить блюдо '{product.name}'? (да/нет): ")
                        if confirm.lower() == 'да':
                            self.excel.delete_product(product.id)
                            restaurant.menu.remove(product)
                            print("Блюдо успешно удалено!")
                elif choice == "4":
                    product = self.select_product(restaurant, "Выберите блюдо для изменения статуса: ")
                    if product:
                        new_status = product.change_status()
                        self.excel.save_product(product)
                        print(f"Статус изменён на {'доступен' if new_status else 'не доступен'}")
                elif choice == "5":
                    return
                else:
                    print("Неверный ввод")

    def main_menu(self):
        while True:
            print("\nГлавное меню:")
            print("1. Добавить ресторан")
            print("2. Редактировать ресторан")
            print("3. Редактировать меню")
            print("4. Просмотреть все рестораны")
            print("5. Просмотреть детали ресторана")
            print("6. Поиск ресторана")
            print("7. Выход")

            choice = input("Выберите действие: ")

            if choice == "1":
                self.create_restaurant()
            elif choice == "2":
                self.edit_restaurant()
            elif choice == "3":
                self.edit_menu()
            elif choice == "4":
                self.show_all_restaurants()
            elif choice == "5":
                self.show_restaurant_details()
            elif choice == "6":
                self.search_restaurant()
            elif choice == "7":
                print("Выход из программы")
                break
            else:
                print("Неверный ввод, попробуйте еще раз")


def main():
    manager = RestaurantManager()

    # Авторизация
    while not manager.current_user:
        if not manager.login():
            retry = input("Хотите попробовать еще раз? (да/нет): ")
            if retry.lower() != 'да':
                return

    # Основное меню
    manager.main_menu()


if __name__ == '__main__':
    main()